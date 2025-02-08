#Tags - ID
#Tags - NAME
#Tags - TAG_NAME
#Tags - LINK
#Tags - CLASS_NAME
#Tags - CSS_SELECTOR
#Tags - XPATH

from openpyxl.worksheet.worksheet import Worksheet
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl import Workbook
from time import sleep
import pandas as pd
import os

# Inicializa o driver do navegador
def main():
    driver = webdriver.Chrome()
    url = 'https://curso-web-scraping.pages.dev/#/exemplo/1'

    driver.get(url)
    nome = driver.find_element(By.ID, 'user').get_property('value')
    proficao = driver.find_element(By.ID, 'role').get_property('value')
    signo = driver.find_element(By.ID, 'zodiac').get_property('value')
    genero = driver.find_element(By.ID, 'gender').get_property('value')
    sleep(1)
    driver.quit()

    # Dados coletados
    dados = {
        'Nome': [nome],
        'Profissão': [proficao],
        'Signo': [signo],
        'Gênero': [genero]
    }

    # Cria um DataFrame com os dados coletados
    df = pd.DataFrame(dados)

    arquivo_csv = 'dados.csv'
    if not os.path.exists(arquivo_csv):
        df.to_csv(arquivo_csv, index=False, mode='w', encoding='utf-8')
    else:
        df.to_csv(arquivo_csv, index=False, mode='a', header=False, encoding='utf-8')

    # Formata o arquivo CSV
    arquivo_formatado_csv = 'dados_formatado.csv'
    with open(arquivo_formatado_csv, 'a', encoding='utf-8') as f:
        for _, linha in df.iterrows():
            for chave, valor in linha.items():
                f.write(f'{chave}: {valor}\n')
            f.write('\n')

    #Arquivo xlsx
    arquivo_xlsx = 'dados2.xlsx'
    if os.path.exists(arquivo_xlsx):
        try:
            workbook = load_workbook(arquivo_xlsx)
            worksheet = workbook.active
        except Exception as e:
            print(f"Erro ao carregar o arquivo Excel: {e}")
            workbook = Workbook()
            worksheet = workbook.active
            # Cria cabeçalhos no caso de erro no carregamento
            worksheet.append(['Nome', 'Profissão', 'Signo', 'Gênero'])
    else:
        # Cria um novo arquivo Excel
        workbook = Workbook()
        worksheet = workbook.active
        # Cria cabeçalhos no arquivo novo
        worksheet.append(['Nome', 'Profissão', 'Signo', 'Gênero'])

    # Adiciona os dados coletados na planilha
    for _, row in df.iterrows():
        worksheet.append(row.tolist())

    workbook.save(arquivo_xlsx)

    # Exibe o conteúdo dos arquivos
    print('\nConteúdo do arquivo formatado CSV:')
    with open(arquivo_formatado_csv, 'r', encoding='utf-8') as f:
        print(f.read())

    print('\nConteúdo do arquivo CSV:')
    with open(arquivo_csv, 'r', encoding='utf-8') as f:
        print(f.read())

    print("\nConteúdo do arquivo Excel:")
    excel_df = pd.read_excel(arquivo_xlsx)  
    print(excel_df)

if __name__ == '__main__':
    main()