from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell # type: ignore
from openpyxl.worksheet.worksheet import Worksheet # type: ignore
import openpyxl


PASTA_RAIZ = Path(__file__).parent
WORKBOOK_PATH = PASTA_RAIZ / 'workbook.xlsx'

workbook = load_workbook(WORKBOOK_PATH)

name_sheet = 'Minha Planilha'
sheet = Worksheet = workbook[name_sheet]

row: tuple[Cell]
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end= '\t')
    print()

# sheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)