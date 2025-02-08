from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet # type: ignore
import openpyxl


PASTA_RAIZ = Path(__file__).parent
WORKBOOK_PATH = PASTA_RAIZ / 'workbook.xlsx'

workbook = Workbook()

name_sheet = 'Minha Planilha'
workbook.create_sheet(name_sheet, 0)
sheet = Worksheet = workbook[name_sheet]

#Remover uma planila
workbook.remove(workbook['Sheet'])

sheet.cell(1, 1, 'Nome')
sheet.cell(1, 2, 'Idade')
sheet.cell(1, 3, 'Nota')

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

#TEM ESSA OPCAO TAMBÉM
# for student in students:
#     sheet.append(student)

for i, students_row in enumerate(students, start=2):
    for j, studants_column in enumerate(students_row, start=1):
        sheet.cell(i, j, studants_column)

workbook.save(WORKBOOK_PATH)